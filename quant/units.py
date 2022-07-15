"""
General utility functions.
"""
import json
import logging
from pathlib import Path
from typing import Callable, Dict, Tuple, Union, Optional
from decimal import Decimal
from math import floor, ceil
import numpy as np
import talib

import openpyxl
from openpyxl.styles import  PatternFill
from openpyxl.utils import get_column_letter
from vnpy.trader.database import get_database

from vnpy.trader.object import Interval


from vnpy.trader.utility import ArrayManager



def get_symbol_overview(symbol, interval=Interval.DAILY):
    #interval = Interval.DAILY
    overview = get_database().get_bar_overview()
    data = None
    for item in overview:
        if item.symbol == symbol.lower() and item.interval == interval:
            data = item
            break
    return data


def bool_color(condition, color='6495ED'):
    if condition:
        return color
    return ''


def excel_xlsx_sheet(workbook, sheet_name, data):
    index = len(data)
    sheet = workbook.create_sheet(sheet_name)
    # sheet = workbook.active
    # sheet.title = sheet_name

    for i in range(0, index):
        for j in range(0, len(data[i])):
            val = data[i][j]
            if i == 0 and not isinstance(val, list):
                val = [val, 'FFA500']

            if isinstance(val, list) and len(val) > 1:
                cell = sheet.cell(row=i + 1, column=j + 1, value=str(val[0]))
                if val[1] != "":
                    fill_fg_color = PatternFill("solid", fgColor=val[1])
                    cell.fill = fill_fg_color
            else:
                sheet.cell(row=i + 1, column=j + 1, value=str(val))
        # if i > 0:
        #     if value[i][5] < 0.0:
        #         sheet['F' + str(i + 1)].fill = fill_fg_red
        #     else:
        #         sheet['F' + str(i + 1)].fill = fill_fg_green
    for i in range(1, sheet.max_column+1):
        width = 10
        if len(data[0]) > i and isinstance(data[0][i-1], list) and len(data[0][i-1]) > 2:
            width = data[0][i-1][2]
        sheet.column_dimensions[get_column_letter(i)].width = width
    sheet.freeze_panes = sheet['B2']

    return sheet


def report_excel_xlsx(save_file, data):
    #index = len(report_trade)
    workbook = openpyxl.Workbook()
    workbook.remove_sheet(workbook.active)

    for ent in data:
        excel_xlsx_sheet(workbook, ent[0], ent[1])

    workbook.save(save_file)


def read_excel_xlsx(path, sheet_name):
    workbook = openpyxl.load_workbook(path)
    # sheet = wb.get_sheet_by_name(sheet_name)这种方式已经弃用，不建议使用
    sheet = workbook[sheet_name]
    for row in sheet.rows:
        for cell in row:
            print(cell.value, "\t", end="")
        print()

# book_name_xlsx = 'xlsx格式测试工作簿.xlsx'
#
# sheet_name_xlsx = 'xlsx格式测试表'

# value3 = [["姓名", "性别", "年龄", "城市", "职业"],
#           ["111", "女", "66", "石家庄", "运维工程师"],
#           ["222", "男", "55", "南京", "饭店老板"],
#           ["333", "女", "27", "苏州", "保安"],]
#
#
# write_excel_xlsx(book_name_xlsx, sheet_name_xlsx, value3)
# read_excel_xlsx(book_name_xlsx, sheet_name_xlsx)
